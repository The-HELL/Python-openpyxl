import pytest
from openpyxl import load_workbook
import sys
import os
sys.path.append("..")
from app.sheet import Sheet


@pytest.fixture
def sheet():
    wb = load_workbook(filename=os.path.join("."+os.sep, "排班导入计划.xlsx"))
    sheet = Sheet(wb[wb.sheetnames[1]])
    return sheet
