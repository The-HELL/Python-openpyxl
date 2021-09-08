from openpyxl import load_workbook
from copy import copy

class Sheet(object):
    def __init__(self, sheet):
        self.sheet = sheet
        self.value_dict = {}

    def get_dates(self, begin_date:str, days:int):
        # 从起始日期读取
        begin_date = begin_date
        dates = []
        dates.append(begin_date)
        i = 0
        column=4
        while i<(days-1):
            date = self.sheet.cell(row=1, column=column).value

            try:
                if begin_date < date:
                    dates.append(date)
                    i += 1
                column += 2
            except TypeError:
                raise TypeError("超过表格中的日期最大限度")

        return dates

    def sheet_append(self, name:list, value, dates:list):
        l_name = copy(name)
        u_row = 2
        u_column = 2
        Len = len(dates)
        i = 0
        username = ""
        # 第一层while，锁定名字
        while (len(l_name) != 0):
            username = self.sheet.cell(row=u_row, column=u_column).value
            if username in l_name:
                l_name.pop(l_name.index(username))

                d_row, d_column = 1, 4

                # 第二层while, 按日期填写value
                while True:
                    date = self.sheet.cell(row=d_row, column=d_column).value
                    if date in dates:
                        # 上午
                        self.sheet.cell(row=(u_row+1), column=(d_column+1)).value = value
                        # 下午
                        self.sheet.cell(row=(u_row+2), column=(d_column+1)).value = value

                        # 将信息添加到self.value_dict
                        self.value_dict[username] = (date, value)

                    # 当date为列表中最大日期时，
                    # 在继续循环取date值时，date会取到None
                    # elif date < dates[0] 比较时，会引发TypeError错误:不能比较'NoneType'和'str'
                    elif date is None:
                        break

                    elif date < dates[0]:
                        pass

                    else:
                        # 超过dates中的最大日期
                        break
                    d_column += 2

            u_row += 4

        return True

    def get_value(self):
        return self.value_dict




if __name__ == "__main__":
    filename = "C:\\Users\\陈陈\\Desktop\\2021-09-01 10_57_10排班导入计划.xlsx"
    wb = load_workbook(filename = filename)
    sheet = Sheet(wb[wb.sheetnames[1]])
    begin_date = "2021.09.02"
    dates = sheet.get_dates(begin_date=begin_date, days=7)


    n = sheet.sheet_append(name=["周立君", "贾春宝"], value="不限", dates=dates)
    if n:
        print("\n\t添加完成\n")
    else:
        print("\n\t添加失败\n")

    wb.save(filename)

    wb.close()
