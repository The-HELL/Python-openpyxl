from openpyxl import load_workbook
from app.sheet import Sheet
import os


def get_name_value(model_filename:str=os.path.join("." + os.sep, "model.xlsx")):
    wb1 = load_workbook(filename=model_filename)
    sheet1 = wb1[wb1.sheetnames[0]]
    name_value = {}

    for i in sheet1.rows:
        # value与其他value不同,则根据其value创建新的key
        if i[1].value not in name_value.keys():
            name_value[i[1].value] = [i[0].value, ]
        else:
            name_value.get(i[1].value).append(i[0].value)

    wb1.close()
    return name_value



def append_xlsx(name_value:dict, filename:str, begin_date:str, days:int):
    wb2 = load_workbook(filename=filename)
    sheet = Sheet(wb2[wb2.sheetnames[1]])
    dates = sheet.get_dates(begin_date=begin_date, days=days)

    try:
        for value in name_value.keys():
            sheet.sheet_append(name=name_value.get(value), value=value, dates=dates)
        wb2.save(filename=filename)
        wb2.close()
        return True
    except EnvironmentError:
        wb2.close()
        return False


if __name__ == "__main__":

    name_value = get_name_value()
    filename = "C:\\Users\\陈陈\\Desktop\\2021-09-01 10_57_10排班导入计划.xlsx"
    begin_date="2021.09.02"
    days = 4

    if append_xlsx(name_value=name_value, filename=filename, begin_date=begin_date, days=days):
        print("\t\n添加完成！\n")
    else:
        print("\t\n添加失败！！\n")
